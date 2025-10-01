import os
import numpy as np
import pandas as pd
from openpyxl.chart import ScatterChart, Reference, Series

# ----------------------------------------------------------------------
# Matplotlibのフォント設定 (この部分はプログラムの動作に影響しません)
# ----------------------------------------------------------------------
try:
    import matplotlib.pyplot as plt
    plt.rcParams["font.family"] = "Times New Roman"
    plt.rcParams['mathtext.fontset'] = 'cm' 
    plt.rcParams["font.size"] = 20
except ImportError:
    pass # Matplotlibがなくてもデータ処理は続行
# ----------------------------------------------------------------------


def extract_upper_branch(h, b):
    """
    ヒステリシスループのデータ(H, B)から、Bの最大点から最小点までを
    「上側の枝（下降曲線）」として抽出します。
    """
    i_min = int(np.argmin(b))
    i_max = int(np.argmax(b))
    
    M = len(h)

    if i_max <= i_min:
        indices = np.arange(i_max, i_min + 1)
    else:
        indices = np.concatenate([np.arange(i_max, M), np.arange(0, i_min + 1)])
        
    return h[indices], b[indices]


def process_single_file(input_filepath, output_filepath):
    """
    単一の入力Excelファイルを処理し、上側ループを抽出して、
    データとグラフを含む新しいExcelファイルとして保存します。
    """
    try:
        df = pd.read_excel(input_filepath, header=None, names=["H", "B"])
        
        H = df["H"].values
        B = df["B"].values

        H_branch, B_branch = extract_upper_branch(H, B)
        
        n_points = len(H_branch)
        Bm_min = float(np.min(B))
        Bm_max = float(np.max(B))

        df_branch = pd.DataFrame({"H": H_branch, "B": B_branch})

        with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
            df_branch.to_excel(writer, sheet_name='Data', index=False)
            
            ws = writer.sheets['Data']
            ws['D2'] = 'Bm_min'
            ws['E2'] = 'Bm_max'
            ws['D3'] = Bm_min
            ws['E3'] = Bm_max
            ws['D5'] = 'Number of Points'
            ws['D6'] = n_points

            if n_points >= 2:
                chart_ws = writer.book.create_sheet('Chart')
                chart = ScatterChart()
                chart.title = f"Upper Branch ({os.path.basename(output_filepath)})"
                chart.x_axis.title = 'H (A/m)'
                chart.y_axis.title = 'B (T)'
                
                xref = Reference(ws, min_col=1, min_row=2, max_row=n_points + 1)
                yref = Reference(ws, min_col=2, min_row=2, max_row=n_points + 1)
                
                series = Series(yref, xref, title_from_data=False)
                series.smooth = True
                series.marker.symbol = 'none'
                chart.series.append(series)
                
                chart_ws.add_chart(chart, 'A1')

        return True

    except Exception as e:
        print(f"   -> ❌ エラー: ファイル {os.path.basename(input_filepath)} の処理中に問題が発生しました: {e}")
        return False


def main():
    """
    メインの実行関数。
    """
    # --- ユーザー設定項目 ---
    mat_name = "50A470"
    frequencies = [20, 50, 100, 200, 400, 600, 800, 1000]
    # -------------------------

    # ★★★ 修正箇所：相対パス設定 ★★★
    # ----------------------------------------------------------------------
    # 1. このスクリプト(.py)があるフォルダの場所を取得します。
    #    (これが GPR_perf フォルダのパスになります)
    base_dir = os.path.dirname(__file__)

    # 2. 入力と出力のベースパスを、上記 base_dir からの相対パスで定義します。
    #    今回は「1.Training Data Folder」の重複がないように注意しました。
    base_input_root = os.path.join(base_dir,  "3.Fourier Transform Correction")
    base_output_root = os.path.join(base_dir, "5.Ascending branch of the hysteresis loop")
    # ----------------------------------------------------------------------

    total_files_processed_all_freqs = 0
    print(f"処理を開始します。対象: {mat_name}, 周波数: {frequencies}")

    for frequency in frequencies:
        print("\n" + "="*70)
        print(f"▶️  周波数: {frequency} Hz の処理を開始します...")
        print("="*70)

        input_dir = os.path.join(base_input_root, mat_name, f"{frequency}Hz")
        output_dir = os.path.join(base_output_root, mat_name, f"{frequency}Hz")

        if not os.path.isdir(input_dir):
            print(f"❌ 入力フォルダが見つかりません: {input_dir}")
            print(f"   周波数 {frequency} Hz の処理をスキップします。")
            continue

        os.makedirs(output_dir, exist_ok=True)

        files_in_dir = os.listdir(input_dir)
        xlsx_files = [f for f in files_in_dir if f.endswith('.xlsx')]

        if not xlsx_files:
            print(f"-> 処理対象のExcelファイル (.xlsx) が見つかりませんでした。")
            continue
        
        print(f"-> '{os.path.basename(input_dir)}' 内の {len(xlsx_files)} 個のファイルを処理します...")
        files_processed_this_freq = 0

        for filename in xlsx_files:
            input_filepath = os.path.join(input_dir, filename)
            output_filepath = os.path.join(output_dir, filename)
            
            if process_single_file(input_filepath, output_filepath):
                files_processed_this_freq += 1
        
        print(f"--- 周波数 {frequency} Hz の処理結果 ---")
        print(f"   処理されたファイル数: {files_processed_this_freq}")
        total_files_processed_all_freqs += files_processed_this_freq

    print("\n" + "="*70)
    print("🎉 全ての周波数の処理が完了しました。")
    print(f"   合計処理ファイル数: {total_files_processed_all_freqs}")
    print("="*70)

if __name__ == '__main__':
    main()