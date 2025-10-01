import os
import numpy as np
import pandas as pd
import math
from openpyxl.chart import ScatterChart, Reference, Series
import matplotlib.pyplot as plt

def find_data_file(base_path, mat, freq, amp):
    """
    指定された振幅のデータファイルを、小数点以下の桁数を考慮して検索する関数。
    """
    amp_str_1f = f"{amp:.1f}"
    amp_str_2f = f"{amp:.2f}"
    
    path_try2 = os.path.join(base_path, mat, f"{freq}Hz", f"Bm{amp_str_2f}hys_{freq}hz.xlsx")
    if os.path.exists(path_try2):
        return path_try2, amp_str_2f
    
    path_try1 = os.path.join(base_path, mat, f"{freq}Hz", f"Bm{amp_str_1f}hys_{freq}hz.xlsx")
    if os.path.exists(path_try1):
        return path_try1, amp_str_1f
        
    return None, None

def load_reduction_map(path):
    """
    reduction_analysis.xlsx の Bm→reducted マップを辞書で返す
    """
    df = pd.read_excel(path, header=0, engine='openpyxl')
    keys = np.round(df['B_m'].values, 2)
    values = df['Reduced_Points'].values
    return dict(zip(keys, values))

# ★★★ 修正箇所１：間引きロジックを「Bmax→Bmin間の線長ベース」に完全変更 ★★★
def reduce_points_by_arclength(h, b, interior_keep):
    """
    H,B の配列からBmax→Bmin間の線長に沿って等間隔に点を削減する。
    始点/終点/Bmax/Bminは必ず保持する。
    """
    M = len(h)
    if M < 4:
        return h, b
        
    # 1. 必ず残す点のインデックスを確保
    i_max = int(np.argmax(b))
    i_min = int(np.argmin(b))
    mandatory_indices = {0, M - 1, i_max, i_min}
    
    # 2. BmaxからBminまでの「上側カーブ」のインデックスリストを作成
    if i_max <= i_min:
        branch_indices = np.arange(i_max, i_min + 1)
    else: # 配列の終端をまたぐ場合
        branch_indices = np.concatenate([np.arange(i_max, M), np.arange(0, i_min + 1)])
    
    # 3. 抽出した上側カーブのデータ
    h_branch = h[branch_indices]
    b_branch = b[branch_indices]

    # 4. 尺度を揃えるために、カーブのHとBをそれぞれ0～1の範囲に正規化
    h_norm = (h_branch - h_branch.min()) / (h_branch.max() - h_branch.min()) if h_branch.max() > h_branch.min() else np.zeros_like(h_branch)
    b_norm = (b_branch - b_branch.min()) / (b_branch.max() - b_branch.min()) if b_branch.max() > b_branch.min() else np.zeros_like(b_branch)

    # 5. 正規化後のデータで、カーブの累積線長を計算
    segment_lengths = np.sqrt(np.diff(h_norm)**2 + np.diff(b_norm)**2)
    cumulative_length = np.concatenate(([0], np.cumsum(segment_lengths)))
    total_length = cumulative_length[-1]

    # 6. 削減後の内部点の数を決定
    k = math.ceil(interior_keep)
    
    sampled_indices_on_branch = set()
    if k > 0 and total_length > 0:
        # 7. 線長に沿って等間隔な「目標距離」のリストを作成
        target_distances = np.linspace(0, total_length, num=k + 2)[1:-1]
        
        # 8. 各目標距離に最も近い点の「カーブ上でのインデックス」を探す
        for dist in target_distances:
            idx = np.argmin(np.abs(cumulative_length - dist))
            sampled_indices_on_branch.add(idx)
    
    # 9. カーブ上でのインデックスを、元の配列全体のインデックスに変換
    original_sampled_indices = {branch_indices[i] for i in sampled_indices_on_branch}

    # 10. 必ず残す点と、サンプリングした点を結合
    final_indices = mandatory_indices.union(original_sampled_indices)

    # 11. インデックスを昇順にソートして、元のデータから抽出して返す
    idxs_sorted = sorted(list(final_indices))
    return h[idxs_sorted], b[idxs_sorted]


def process_amplitude(amplitude, freq, mat_name, input_base, output_base, reduction_map):
    """
    指定された単一の振幅と周波数に対して処理を実行する関数。
    """
    in_path, amp_str_found = find_data_file(input_base, mat_name, freq, amplitude)
    
    if in_path is None:
        return None, None
        
    try:
        df = pd.read_excel(in_path, header=None, usecols=[0, 1])
        df.columns = ['H', 'B']
        df['H'] = pd.to_numeric(df['H'], errors='coerce')
        df['B'] = pd.to_numeric(df['B'], errors='coerce')
        df.dropna(inplace=True)
    except Exception as e:
        print(f"ファイル読み込みまたはデータクレンジング中のエラー: {in_path}, エラー: {e}")
        return None, None
    
    if df.empty:
        print(f"警告: ファイル {os.path.basename(in_path)} は有効な数値データを含んでいないため、スキップします。")
        return None, None

    H = df["H"].values
    B = df["B"].values

    key = round(amplitude, 2)
    interior_keep = reduction_map.get(key)
    
    if interior_keep is None:
        print(f"警告: 振幅 {key:.2f}T の削減数がマップにありません。スキップします。")
        return None, None

    H_red, B_red = reduce_points_by_arclength(H, B, interior_keep) # 新しい関数を呼び出し
    total_points = len(H_red)

    out_dir = os.path.join(output_base, mat_name, str(freq))
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"Bm{amp_str_found}hys_{freq}hz_reduct.xlsx")

    df_red = pd.DataFrame({'H': H_red, 'B': B_red})
    # ... (Excelへの書き込み処理は変更なし) ...
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_red.to_excel(writer, sheet_name='Data', index=False)
        ws = writer.sheets['Data']
        Bm_max = float(np.max(B)) if len(B) > 0 else 0
        Bm_min = float(np.min(B)) if len(B) > 0 else 0
        ws['E2'] = 'Bm_max'
        ws['E3'] = Bm_max
        ws['E4'] = 'Bm_min'
        ws['E5'] = Bm_min
        ws['E7'] = 'Data Count'
        ws['E8'] = total_points

        if total_points >= 2:
            chart_ws = writer.book.create_sheet('Chart')
            chart = ScatterChart()
            chart.title = f"Reduced Branch ({amp_str_found}T)"
            chart.x_axis.title = 'H'
            chart.y_axis.title = 'B'
            xref = Reference(ws, min_col=1, min_row=2, max_row=total_points+1)
            yref = Reference(ws, min_col=2, min_row=2, max_row=total_points+1)
            series = Series(yref, xref, title_from_data=False)
            series.smooth = True
            series.marker.symbol = 'none'
            chart.series.append(series)
            chart_ws.add_chart(chart, 'A1')
            
    print(f"   -> 保存完了: {os.path.basename(out_path)}")
    return H, B, H_red, B_red, f"Bm={amp_str_found}T, {freq}Hz"


def main():
    # --- ユーザー設定箇所 ---
    mat_name = "50A470"
    amplitudes = np.round(np.arange(0.05, 2.0 + 1e-8, 0.05), 2)
    frequencies = [20, 50, 100, 200, 400, 600, 800, 1000]
    # --- 設定ここまで ---

    # --- 相対パス設定 ---
    # 1. このスクリプト(.py)があるフォルダの場所を取得
    script_dir = os.path.dirname(__file__)
    # 2. スクリプトが置かれている 'src' フォルダの親フォルダを基準パスとします。
    base_dir = os.path.dirname(script_dir)
    reduction_map_path = os.path.join(base_dir, "assets", "4.Reduction Point Determination Process", "reduction_analysis.xlsx")
    input_base = os.path.join(base_dir, "assets", "5.Ascending branch of the hysteresis loop")
    output_base = os.path.join(base_dir, "assets", "6.2 Downsampling at s")
    
    print(f"削減マップを読み込んでいます: {reduction_map_path}")
    try:
        reduction_map = load_reduction_map(reduction_map_path)
    except FileNotFoundError:
        print(f"❌ エラー: 削減マップファイルが見つかりません。パスを確認してください。")
        return
    except Exception as e:
        print(f"❌ エラー: 削減マップファイルの読み込み中に予期せぬエラーが発生しました: {e}")
        return

    print(f"処理を開始します。対象周波数: {frequencies}")
    
    last_plot_data = None
    
    for freq in frequencies:
        print("\n" + "="*70)
        print(f"▶️  周波数: {freq} Hz の処理を開始します...")
        print("="*70)
        
        for amp in amplitudes:
            plot_data = process_amplitude(amp, freq, mat_name, input_base, output_base, reduction_map)
            if plot_data and plot_data[0] is not None:
                last_plot_data = plot_data
            
        print(f"-> {freq}Hz の処理が完了しました。")

    # --- 処理完了後に結果をプロット ---
    if last_plot_data:
        print("\n" + "="*70)
        print("▶️  最終処理結果のサンプルをプロットします...")
        
        h_orig, b_orig, h_red, b_red, title_info = last_plot_data
        
        plt.figure(figsize=(10, 8))
        plt.plot(h_orig, b_orig, color='lightgray', linestyle='-', label=f'Original ({len(h_orig)} pts)')
        plt.plot(h_red, b_red, 'o-', color='royalblue', label=f'Downsampled ({len(h_red)} pts)')
        
        b_max_idx_red = np.argmax(b_red)
        b_min_idx_red = np.argmin(b_red)
        plt.scatter([h_red[b_max_idx_red], h_red[b_min_idx_red]],
                    [b_red[b_max_idx_red], b_red[b_min_idx_red]],
                    color='red', s=120, zorder=5, edgecolors='white', label='Bmax/Bmin Points')

        plt.title(f'Downsampling Result (by Arc Length)\n{title_info}')
        plt.xlabel('H (A/m)')
        plt.ylabel('B (T)')
        plt.legend()
        plt.grid(True)
        plt.show()
    else:
        print("\n⚠️ 処理されたデータがないため、最終結果のプロットはスキップされました。")


if __name__ == '__main__':
    main()