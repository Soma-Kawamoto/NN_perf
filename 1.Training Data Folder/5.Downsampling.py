import os
import numpy as np
import pandas as pd
import math
from openpyxl.chart import ScatterChart, Reference, Series

def find_data_file(base_path, mat, freq, amp):
    """
    指定された振幅のデータファイルを、小数点以下の桁数を考慮して検索する関数。
    """
    amp_str_1f = f"{amp:.1f}"
    amp_str_2f = f"{amp:.2f}"
    
    path_try2 = os.path.join(base_path, mat, f"{freq}Hz", f"Bm{amp_str_2f}hys_{freq}hz.xlsx")
    if os.path.exists(path_try2):
        return path_try2, amp_str_2f
    
    path_try1 = os.path.join(base_path, mat, f"{freq}Hz", f"Bm{amp_str_1f}hys_{freq}hz.xlsx")
    if os.path.exists(path_try1):
        return path_try1, amp_str_1f
        
    return None, None

def load_reduction_map(path):
    """
    reduction_analysis.xlsx の Bm→reducted マップを辞書で返す
    """
    df = pd.read_excel(path, header=0, engine='openpyxl')
    keys = np.round(df['B_m'].values, 2)
    values = df['Reduced_Points'].values
    return dict(zip(keys, values))

def reduce_points(h, b, interior_keep):
    """
    H,B の配列から最大/最小点、先頭/末尾を残し、
    間の interior_keep 点を等間隔に選択して削減する。
    """
    M = len(h)
    if M < 4:
        return h, b
        
    idxs = {0, M-1, int(np.argmax(b)), int(np.argmin(b))}
    
    k = math.ceil(interior_keep)
    k = min(k, M - len(idxs))
    
    if k > 0:
        available_indices = [i for i in range(1, M - 1) if i not in idxs]
        
        if len(available_indices) > k:
            select_indices = np.linspace(0, len(available_indices) - 1, num=k, endpoint=True)
            select_indices = np.round(select_indices).astype(int)
            select_indices = np.unique(select_indices)
            chosen_original_indices = np.array(available_indices)[select_indices]
            idxs.update(chosen_original_indices.tolist())
        else:
            idxs.update(available_indices)

    idxs_sorted = sorted(list(idxs))
    return h[idxs_sorted], b[idxs_sorted]

def process_amplitude(amplitude, freq, mat_name, input_base, output_base, reduction_map):
    """
    指定された単一の振幅と周波数に対して処理を実行する関数。
    """
    in_path, amp_str_found = find_data_file(input_base, mat_name, freq, amplitude)
    
    if in_path is None:
        return
        
    try:
        df = pd.read_excel(in_path, header=None, usecols=[0, 1])
        df.columns = ['H', 'B']
        df['H'] = pd.to_numeric(df['H'], errors='coerce')
        df['B'] = pd.to_numeric(df['B'], errors='coerce')
        df.dropna(inplace=True)

    except Exception as e:
        print(f"ファイル読み込みまたはデータクレンジング中のエラー: {in_path}, エラー: {e}")
        return
    
    if df.empty:
        print(f"警告: ファイル {os.path.basename(in_path)} は有効な数値データを含んでいないため、スキップします。")
        return

    H = df["H"].values
    B = df["B"].values

    key = round(amplitude, 2)
    interior_keep = reduction_map.get(key)
    
    if interior_keep is None:
        print(f"警告: 振幅 {key:.2f}T の削減数がマップにありません。スキップします。")
        return

    H_red, B_red = reduce_points(H, B, interior_keep)
    total_points = len(H_red)

    out_dir = os.path.join(output_base, mat_name, str(freq))
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"Bm{amp_str_found}hys_{freq}hz_reduct.xlsx")

    df_red = pd.DataFrame({'H': H_red, 'B': B_red})
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_red.to_excel(writer, sheet_name='Data', index=False)
        ws = writer.sheets['Data']
        Bm_max = float(np.max(B)) if len(B) > 0 else 0
        Bm_min = float(np.min(B)) if len(B) > 0 else 0
        ws['E2'] = 'Bm_max'
        ws['E3'] = Bm_max
        ws['E4'] = 'Bm_min'
        ws['E5'] = Bm_min
        ws['E7'] = 'Data Count'
        ws['E8'] = total_points

        if total_points >= 2:
            chart_ws = writer.book.create_sheet('Chart')
            chart = ScatterChart()
            chart.title = f"Reduced Branch ({amp_str_found}T)"
            chart.x_axis.title = 'H'
            chart.y_axis.title = 'B'
            xref = Reference(ws, min_col=1, min_row=2, max_row=total_points+1)
            yref = Reference(ws, min_col=2, min_row=2, max_row=total_points+1)
            series = Series(yref, xref, title_from_data=False)
            series.smooth = True
            series.marker.symbol = 'none'
            chart.series.append(series)
            chart_ws.add_chart(chart, 'A1')

    print(f"   -> 保存完了: {os.path.basename(out_path)}")

def main():
    # --- ユーザー設定箇所 ---
    mat_name = "50A470"
    amplitudes = np.round(np.arange(0.05, 2.0 + 1e-8, 0.05), 2)
    frequencies = [20, 50, 100, 200, 400, 600, 800, 1000]
    # --- 設定ここまで ---

    # ★★★ 修正箇所：相対パス設定 ★★★
    # ----------------------------------------------------------------------
    # 1. このスクリプト(.py)がある GPR_perf フォルダの場所を取得
    base_dir = os.path.dirname(__file__)

    # 2. base_dir からの相対的な位置で各パスを定義
    reduction_map_path = os.path.join(base_dir, "4.Reduction Point Determination Process", "reduction_analysis.xlsx")
    input_base = os.path.join(base_dir,  "5.Ascending branch of the hysteresis loop")
    output_base = os.path.join(base_dir, "6.Downsampling")
    # ----------------------------------------------------------------------

    print(f"削減マップを読み込んでいます: {reduction_map_path}")
    try:
        reduction_map = load_reduction_map(reduction_map_path)
    except FileNotFoundError:
        print(f"❌ エラー: 削減マップファイルが見つかりません。パスを確認してください。")
        return
    except KeyError as e:
        print(f"❌ エラー: 削減マップファイルのヘッダー名が正しくありません。'{e}' という列が見つかりません。")
        print(f"   ファイル '{os.path.basename(reduction_map_path)}' のヘッダーが 'B_m' と 'Reduced_Points' になっているか確認してください。")
        return

    print(f"処理を開始します。対象周波数: {frequencies}")
    
    for freq in frequencies:
        print("\n" + "="*70)
        print(f"▶️  周波数: {freq} Hz の処理を開始します...")
        print("="*70)
        
        for amp in amplitudes:
            process_amplitude(amp, freq, mat_name, input_base, output_base, reduction_map)
            
        print(f"-> {freq}Hz の処理が完了しました。")

if __name__ == '__main__':
    main()