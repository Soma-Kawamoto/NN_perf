import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
import matplotlib.pyplot as plt

# ----------------------------------------------------------------------
# ▼▼▼ 変更点 ▼▼▼
# ----------------------------------------------------------------------
plt.rcParams["font.family"] = "Times New Roman"
# 数式部分のフォントもTimes New Roman系に設定
plt.rcParams['mathtext.fontset'] = 'cm' 
plt.rcParams["font.size"] = 20
# ----------------------------------------------------------------------

def fit_quadratic(points):
    """
    3点を通る二次関数 f(x)=a x^2 + b x + c の係数を求める。
    """
    (x1, y1), (x2, y2), (x3, y3) = points
    A = np.array([
        [x1**2, x1, 1],
        [x2**2, x2, 1],
        [x3**2, x3, 1]
    ], dtype=float)
    y = np.array([y1, y2, y3], dtype=float)
    return np.linalg.solve(A, y)

def make_table(a, b, c, bm_values):
    """
    Bm の配列に対して、reducted = a*Bm^2 + b*Bm + c を計算する。
    """
    # BmをX軸、reductedをY軸とするため、列名を変更
    reducted = a * bm_values**2 + b * bm_values + c
    return pd.DataFrame({
        'B_m': bm_values,
        'Reduced_Points': reducted
    })

def plot_graph(ax, a, b, c, points, df):
    """
    計算した二次関数と入力点をプロットする。(pltからaxに変更)
    """
    # 凡例用の数式ラベル
    equation = f'$y = {a:.3f}x^2 + {b:.3f}x + {c:.3f}$'
    ax.plot(df['B_m'], df['Reduced_Points'], label=equation, color='royalblue')
    
    # 入力点をプロット
    x_points = [p[0] for p in points]
    y_points = [p[1] for p in points]
    ax.scatter(x_points, y_points, color='red', zorder=5, s=80, label='Input Points')

    # Bmの表記を修正し、軸ラベルを英語に
    ax.set_xlabel('$B_{\mathrm{m}}$ [T]', fontsize=14)
    ax.set_ylabel('Number of Reduced Points', fontsize=14)
    ax.grid(True, which='both', linestyle='--', linewidth=0.5)
    ax.legend(fontsize=12)

def main():
    # ──── ユーザー設定箇所 ────
    points        = [(0.1, 15), (1.0, 60), (1.8, 150)]
    bm_start      = 0.0
    bm_end        = 2.0
    primary_step  = 0.05
    base_filename = "reduction_analysis" # 保存するファイル名のベース

    # ★★★ 修正箇所：相対パス設定 ★★★
    # ----------------------------------------------------------------------
    # 1. このスクリプト(.py)がある GPR_perf フォルダの場所を取得
    base_dir = os.path.dirname(__file__)

    # 2. base_dir からの相対的な位置で出力先フォルダを指定
    output_dir = os.path.join(base_dir, "4.Reduction Point Determination Process")
    # ----------------------------------------------------------------------
    
    # ExcelファイルとPNGファイルのフルパスを生成
    excel_output_path = os.path.join(output_dir, f"{base_filename}.xlsx")
    plot_output_path = os.path.join(output_dir, f"{base_filename}.png")
    
    # フォルダが存在しない場合は作成
    os.makedirs(output_dir, exist_ok=True)
    
    a, b, c = fit_quadratic(points)
    print(f"Calculated quadratic function: f(x) = {a:.6g} x^2 + {b:.6g} x + {c:.6g}")

    bm_values_plot = np.arange(bm_start, bm_end + 1e-8, 0.01)
    df_plot = make_table(a, b, c, bm_values_plot)
    print("\nGenerated data (first 5 rows):")
    print(df_plot.head())

    fig, ax = plt.subplots(figsize=(10, 7))
    plot_graph(ax, a, b, c, points, df_plot)

    title_text = 'Visualization of Reduction Method (Quadratic Fit)'
    fig.suptitle(title_text, y=0.02, fontsize=16)

    plt.tight_layout(rect=[0, 0.05, 1, 1])
    plt.savefig(plot_output_path)
    print(f"\nGraph saved to: {plot_output_path}")
    plt.xlim(0, 2.0)
    plt.show()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(['B_m', 'Reduced_Points'])

    df_for_excel = make_table(a, b, c, np.arange(bm_start, bm_end + 1e-8, primary_step))
    for row in df_for_excel.itertuples(index=False):
        ws.append(list(row))

    ws["E2"] = "Input points:"
    ws["E3"] = f"({points[0][0]}, {points[0][1]})"
    ws["E4"] = f"({points[1][0]}, {points[1][1]})"
    ws["E5"] = f"({points[2][0]}, {points[2][1]})"

    wb.save(excel_output_path)
    print(f"Excel file saved to: {excel_output_path}")

    print("\n--- Additional Information ---")
    steps_to_calculate = [0.05, 0.1]

    for step_val in steps_to_calculate:
        bm_values = np.arange(step_val, bm_end + 1e-8, step_val)
        reducted_values = a * bm_values**2 + b * bm_values + c
        ceiled_values = np.ceil(reducted_values)
        total_sum = np.sum(ceiled_values)
        total_sum_3 = total_sum**3

        print(f"\nFor step = {step_val} (from Bm={step_val} to {bm_end}):")
        print(f"   Sum of ceiling integer values (Reduced_Points): {int(total_sum)}")
        print(f"   Cube of the sum: {int(total_sum_3)}")

if __name__ == '__main__':
    main()